import openpyxl
import numpy as np

FIG2_PATH = "figure2.xlsx"   # or your actual path

FIG2_PATH = "figure2.xlsx"

def extract_fig2_panel_experimental(fig2_xlsx_path: str, panel_label: str = "panel d"):
    wb = openpyxl.load_workbook(fig2_xlsx_path, data_only=True)
    ws = wb["Fig2"]
    start_col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(1, c).value == panel_label:
            start_col = c
            break
    if start_col is None:
        raise ValueError(f"Panel label '{panel_label}' not found in row 1.")

    exp_col = None
    for c in range(start_col, min(ws.max_column + 1, start_col + 40)):
        v = ws.cell(2, c).value
        if v is not None and "experimental curves" in str(v):
            exp_col = c
            break
    if exp_col is None:
        raise ValueError(f"Could not find 'experimental curves (dashed)' header in panel '{panel_label}'.")

    rho_col = start_col                     # row 3 says "edge density" here
    beta_cols = [exp_col, exp_col + 1, exp_col + 2]  # row 3: red/green/blue

    r0 = 4
    r = r0
    while r <= ws.max_row and ws.cell(r, rho_col).value is not None:
        r += 1
    r1 = r - 1
    rho = np.array([ws.cell(rr, rho_col).value for rr in range(r0, r1 + 1)], dtype=float)
    beta_exp = np.zeros((len(rho), 3), dtype=float)
    for j, cc in enumerate(beta_cols):
        beta_exp[:, j] = [ws.cell(rr, cc).value for rr in range(r0, r1 + 1)]

    meta = {
        "panel_label": panel_label,
        "start_col": start_col,
        "exp_col": exp_col,
        "rows": (r0, r1),
        "rho_col": rho_col,
        "beta_cols": beta_cols,
        "color_to_betti": {"red": "beta1", "green": "beta2", "blue": "beta3"},
    }
    return rho, beta_exp, meta

rhos, beta_emp, meta = extract_fig2_panel_experimental(FIG2_PATH, panel_label="panel d")

from test import simulate_many_corr

# rhos_out, summary, raw = simulate_many_corr(
#     B=50,
#     N=41,
#     Rmax=15.5,
#     Rtrack=15.5,
#     sigma=2.0,
#     T=800,
#     rhos=rhos,
#     seed0=0
# )


# I1_sim = np.trapz(raw[1], x=rhos, axis=1)
# I2_sim = np.trapz(raw[2], x=rhos, axis=1)
# I3_sim = np.trapz(raw[3], x=rhos, axis=1)


sigmas = [0.6, 0.9, 1.3, 1.9, 2.7, 3.8, 5.3]
Rtracks = [8.0, 12.0, 15.5]

B_sweep = 50  # fast
results = []

for Rtrack in Rtracks:
    for sigma in sigmas:
        rhos_out, summary, raw = simulate_many_corr(
            B=B_sweep, N=41, Rmax=15.5, Rtrack=Rtrack, sigma=sigma, T=800, rhos=rhos, seed0=0
        )

        # integrated betti (sim mean)
        I_sim = {}
        for k in [1,2,3]:
            mean_k, _ = summary[k]
            I_sim[k] = float(np.trapz(mean_k, x=rhos_out))

        # integrated betti (empirical)
        I_emp = {
            1: float(np.trapz(beta_emp[:,0], x=rhos)),
            2: float(np.trapz(beta_emp[:,1], x=rhos)),
            3: float(np.trapz(beta_emp[:,2], x=rhos)),
        }

        score = sum(abs(I_sim[k] - I_emp[k]) for k in [1,2,3])

        results.append((score, Rtrack, sigma, I_sim[1], I_sim[2], I_sim[3]))

results.sort(key=lambda x: x[0])
print("Top 10 (lower is better):")
for row in results[:10]:
    print(row)
