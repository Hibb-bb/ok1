import numpy as np
import openpyxl


from simulate import simulate_many

FIG2_PATH = "figure2.xlsx"

def extract_fig2_panel_experimental(fig2_xlsx_path: str, panel_label: str = "panel d"):
    wb = openpyxl.load_workbook(fig2_xlsx_path, data_only=True)
    ws = wb["Fig2"]
    start_col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(1, c).value == panel_label:
            start_col = c
            break
    if start_col is None:
        raise ValueError(f"Panel label '{panel_label}' not found in row 1.")

    exp_col = None
    for c in range(start_col, min(ws.max_column + 1, start_col + 40)):
        v = ws.cell(2, c).value
        if v is not None and "experimental curves" in str(v):
            exp_col = c
            break
    if exp_col is None:
        raise ValueError(f"Could not find 'experimental curves (dashed)' header in panel '{panel_label}'.")

    rho_col = start_col                     # row 3 says "edge density" here
    beta_cols = [exp_col, exp_col + 1, exp_col + 2]  # row 3: red/green/blue

    r0 = 4
    r = r0
    while r <= ws.max_row and ws.cell(r, rho_col).value is not None:
        r += 1
    r1 = r - 1
    rho = np.array([ws.cell(rr, rho_col).value for rr in range(r0, r1 + 1)], dtype=float)
    beta_exp = np.zeros((len(rho), 3), dtype=float)
    for j, cc in enumerate(beta_cols):
        beta_exp[:, j] = [ws.cell(rr, cc).value for rr in range(r0, r1 + 1)]

    meta = {
        "panel_label": panel_label,
        "start_col": start_col,
        "exp_col": exp_col,
        "rows": (r0, r1),
        "rho_col": rho_col,
        "beta_cols": beta_cols,
        "color_to_betti": {"red": "beta1", "green": "beta2", "blue": "beta3"},
    }
    return rho, beta_exp, meta

rho, beta_exp, meta = extract_fig2_panel_experimental(FIG2_PATH, panel_label="panel d")
print(meta)
print("rho shape:", rho.shape, "beta_exp shape:", beta_exp.shape)
print("first rows:", np.c_[rho[:5], beta_exp[:5]])


# import matplotlib.pyplot as plt

# plt.plot(rho, beta_exp[:, 0], label="β1 (exp)")
# plt.plot(rho, beta_exp[:, 1], label="β2 (exp)")
# plt.plot(rho, beta_exp[:, 2], label="β3 (exp)")
# plt.xlabel("edge density ρ")
# plt.ylabel("Betti number")
# plt.legend()
# plt.savefig("./a.png")


# if __name__ == "__main__":
N = 41
Rmax = 15.5
B = 300
seed0 = 0

rho, beta_exp, meta = extract_fig2_panel_experimental(FIG2_PATH, panel_label="panel d")
print(meta)
print("rho shape:", rho.shape, "beta_exp shape:", beta_exp.shape)
print("first rows:", np.c_[rho[:5], beta_exp[:5]])

# Run simulations
rhos_out, summary = simulate_many(B=B, N=N, Rmax=Rmax, rhos=rhos, seed0=seed0)

# Plot mean ± sd for β1–β3
for k in [1, 2, 3]:
    mean_k, sd_k = summary[k]
    plt.plot(rhos_out, mean_k, label=f"sim β{k} mean")
    plt.fill_between(rhos_out, mean_k - sd_k, mean_k + sd_k, alpha=0.2)

plt.xlabel("edge density ρ")
plt.ylabel("Betti number")
plt.title(f"Hyperbolic simulation (N={N}, d=3, Rmax={Rmax}, B={B})")
plt.legend()
plt.show()