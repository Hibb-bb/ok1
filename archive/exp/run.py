import numpy as np
import openpyxl
import matplotlib.pyplot as plt

# If you have these in simulate.py, import them directly:
# from simulate import simulate_one_betti_curve
# Otherwise, keep simulate_many and also import simulate_one_betti_curve from the same file.
from simulate import simulate_one_betti_curve  # <-- preferred for stats

try:
    from scipy.stats import norm
    _HAS_SCIPY = True
except Exception:
    _HAS_SCIPY = False

FIG2_PATH = "figure2.xlsx"  # set to your local path


def _find_panel_start_col(ws, panel_label: str):
    for c in range(1, ws.max_column + 1):
        if ws.cell(1, c).value == panel_label:
            return c
    raise ValueError(f"Could not find '{panel_label}' in row 1.")


def extract_fig2_panel_d_rho(fig2_xlsx_path: str, panel_label: str = "panel d"):
    wb = openpyxl.load_workbook(fig2_xlsx_path, data_only=True)
    ws = wb["Fig2"]

    start_col = _find_panel_start_col(ws, panel_label)
    rho_col = start_col

    r0 = 4
    r = r0
    while r <= ws.max_row and ws.cell(r, rho_col).value is not None:
        r += 1
    r1 = r - 1

    rho = np.array([ws.cell(rr, rho_col).value for rr in range(r0, r1 + 1)], dtype=float)
    return rho


def extract_fig2_panel_experimental_betti(fig2_xlsx_path: str, panel_label: str = "panel d"):
    """
    Returns:
      rho: (T,)
      beta_exp: (T,3) with columns [β1, β2, β3] in the paper's color order (red, green, blue).
    """
    wb = openpyxl.load_workbook(fig2_xlsx_path, data_only=True)
    ws = wb["Fig2"]

    start_col = _find_panel_start_col(ws, panel_label)

    # Find the "experimental curves (dashed)" header in row 2 within this panel block
    exp_col = None
    for c in range(start_col, min(ws.max_column + 1, start_col + 60)):
        v = ws.cell(2, c).value
        if v is not None and "experimental curves" in str(v):
            exp_col = c
            break
    if exp_col is None:
        raise ValueError(f"Could not find 'experimental curves (dashed)' in panel '{panel_label}'.")

    rho_col = start_col
    beta_cols = [exp_col, exp_col + 1, exp_col + 2]  # red/green/blue = β1/β2/β3

    # Data rows
    r0 = 4
    r = r0
    while r <= ws.max_row and ws.cell(r, rho_col).value is not None:
        r += 1
    r1 = r - 1

    rho = np.array([ws.cell(rr, rho_col).value for rr in range(r0, r1 + 1)], dtype=float)
    beta_exp = np.stack(
        [
            np.array([ws.cell(rr, c).value for rr in range(r0, r1 + 1)], dtype=float)
            for c in beta_cols
        ],
        axis=1,
    )
    return rho, beta_exp


def integrate_betti(rho: np.ndarray, beta: np.ndarray) -> float:
    """
    Trapezoidal integral of a Betti curve over rho.
    beta: shape (T,)
    """
    return float(np.trapz(beta, x=rho))


if __name__ == "__main__":
    # ---- Zhang et al. linear track parameters ----
    N = 41
    Rmax = 15.5
    B = 300
    seed0 = 0
    panel = "panel d"

    # 1) Load empirical curves (panel d)
    rho_emp, beta_emp = extract_fig2_panel_experimental_betti(FIG2_PATH, panel_label=panel)

    # 2) Simulate B replicates on the same rho grid and collect per-replicate curves
    #    beta_sims[k] will be array shape (B, T)
    T = len(rho_emp)
    beta_sims = {1: np.zeros((B, T), dtype=float),
                 2: np.zeros((B, T), dtype=float),
                 3: np.zeros((B, T), dtype=float)}

    for b in range(B):
        _, bettis = simulate_one_betti_curve(N=N, Rmax=Rmax, rhos=rho_emp, seed=seed0 + b)
        beta_sims[1][b, :] = bettis[1]
        beta_sims[2][b, :] = bettis[2]
        beta_sims[3][b, :] = bettis[3]

    # 3) Compute mean ± sd for plotting
    summary = {k: (beta_sims[k].mean(axis=0), beta_sims[k].std(axis=0)) for k in [1, 2, 3]}

    # 4) Plot: simulated mean±sd + empirical dashed
    for k in [1, 2, 3]:
        mean_k, sd_k = summary[k]
        plt.plot(rho_emp, mean_k, label=f"sim β{k} mean")
        plt.fill_between(rho_emp, mean_k - sd_k, mean_k + sd_k, alpha=0.2)
        plt.plot(rho_emp, beta_emp[:, k - 1], linestyle="--", label=f"empirical β{k}")

    plt.xlabel("edge density ρ")
    plt.ylabel("number of cycles")  # <-- what you asked for
    plt.title(f"Hyperbolic simulation vs data (panel d; N={N}, d=3, Rmax={Rmax}, B={B})")
    plt.legend()
    plt.tight_layout()
    plt.savefig("b.png", dpi=300)

    # 5) Integrated cycle counts and statistical test
    print("\nIntegrated cycle counts I_k = ∫ beta_k(ρ) dρ (trapezoid on rho grid)\n")

    for k in [1, 2, 3]:
        # empirical integrated
        I_emp = integrate_betti(rho_emp, beta_emp[:, k - 1])

        # simulated distribution
        I_sim = np.array([integrate_betti(rho_emp, beta_sims[k][b, :]) for b in range(B)], dtype=float)

        mu = I_sim.mean()
        sd = I_sim.std(ddof=1)

        # z-score vs simulated distribution
        z = (I_emp - mu) / sd if sd > 0 else np.nan

        # nonparametric two-sided p-value from simulation replicates
        # (how often is a simulated integrated value at least as extreme as empirical, relative to sim mean)
        sim_dev = np.abs(I_sim - mu)
        emp_dev = abs(I_emp - mu)
        p_perm = (np.sum(sim_dev >= emp_dev) + 1.0) / (B + 1.0)

        # optional normal-approx p-value
        if _HAS_SCIPY and np.isfinite(z):
            p_norm = 2.0 * (1.0 - norm.cdf(abs(z)))
        else:
            p_norm = None

        print(f"β{k}: I_emp={I_emp:.4f},  I_sim_mean={mu:.4f},  I_sim_sd={sd:.4f},  z={z:.3f},  p_sim(two-sided)={p_perm:.4g}"
              + (f",  p_norm={p_norm:.4g}" if p_norm is not None else ""))
